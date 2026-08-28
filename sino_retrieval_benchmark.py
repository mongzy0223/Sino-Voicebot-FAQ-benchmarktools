#!/usr/bin/env python3
"""
Sino Voicebot FAQ retrieval benchmark tool.

Calls the setsail multi-graph "text_retrieval_test" endpoint for a batch of
test queries, matches the retrieved FAQ_ID(s) against expected ground-truth
answers, and writes a scored results workbook (per-query detail + summary
metrics, overall and per FAQ category/path).

Usage:
    python3 sino_retrieval_benchmark.py --testcases testcases.xlsx --output results.xlsx

    # Inspect the raw API response shape before running a full batch:
    python3 sino_retrieval_benchmark.py --probe-query "How do I apply for a mortgage?" --probe-path FAQ_Mortgage

See README_benchmark.md for full documentation.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import requests

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing dependency 'openpyxl'. Install with: pip install openpyxl", file=sys.stderr)
    raise

DEFAULT_API_URL = "https://dev.setsailapi.com/nlp/llmGraph/graph/multi/run"
DEFAULT_PROJECT_ID = "public"
DEFAULT_CMS_PROJECT_ID = "sinoexvoicebot-52ynuou"
DEFAULT_GRAPH_NAME = "text_retrieval_test"
ID_KEY_NAME = "FAQ_ID"
DISPLAY_COLUMNS = ["FAQ_ID", "FAQ"]
RETRIEVED_TEXT_KEY_NAME = "FAQ_answer_listening_friendly"
VALID_PATHS = ("FAQ_Mortgage", "FAQ_Leasing")

TESTCASE_QUERY_COLS = ["Query", "Question", "query", "question", "Testing Query"]
TESTCASE_EXPECTED_COLS = [
    "Expected_FAQ_ID", "Expected FAQ_ID", "FAQ_ID", "expected_faq_id",
    "FAQ Reference No.", "FAQ Reference No",
]
TESTCASE_PATH_COLS = ["Path", "Category", "path"]
TESTCASE_LANG_COLS = ["Language", "Lang", "query_lang", "language"]
TESTCASE_NOTES_COLS = ["Notes", "notes"]

# Lets a "Category" column (e.g. Sino's own benchmark template, which uses
# "Leasing"/"Mortgage" rather than the API's FAQ_Leasing/FAQ_Mortgage path
# names) resolve straight to a valid Path. Extend this if new categories
# are added on the CMS side.
CATEGORY_TO_PATH = {
    "leasing": "FAQ_Leasing",
    "mortgage": "FAQ_Mortgage",
}

# When a workbook has no single "Test Cases" sheet, each sheet is treated as
# its own set of test cases and its query_lang is inferred from a language
# suffix in the sheet name (e.g. "Benchmark Template cn" -> sc). Add
# entries here if other sheet-name suffixes/languages are used.
SHEET_NAME_LANG_SUFFIXES = {
    "zh": "zh",
    "en": "en",
    "cn": "sc",
}

# Rows whose query starts with an "EG:" / "EG -" marker are template
# instructional examples, not real test cases, and are skipped.
EG_PREFIX_RE = re.compile(r"^\s*eg\s*[:\-]", re.IGNORECASE)


# --------------------------------------------------------------------------- #
# Data model
# --------------------------------------------------------------------------- #

@dataclass
class TestCase:
    row_num: int
    query: str
    expected_ids: list[str]
    path: str
    lang: str
    notes: str = ""
    sheet: str = ""


@dataclass
class TestResult:
    case: TestCase
    retrieved: list[dict[str, Any]] = field(default_factory=list)
    rank: Optional[int] = None
    latency_ms: Optional[float] = None
    http_status: Optional[int] = None
    error: str = ""
    raw_response: Any = None

    @property
    def top1_correct(self) -> bool:
        return self.rank == 1

    @property
    def reciprocal_rank(self) -> float:
        return 1.0 / self.rank if self.rank else 0.0

    def hit_at(self, k: int) -> bool:
        return self.rank is not None and self.rank <= k

    def retrieved_id(self, position: int) -> str:
        if position < len(self.retrieved):
            return str(self.retrieved[position].get(ID_KEY_NAME, ""))
        return ""

    def retrieved_response(self, position: int) -> str:
        if position < len(self.retrieved):
            return str(self.retrieved[position].get(RETRIEVED_TEXT_KEY_NAME, ""))
        return ""

    @property
    def matched_id(self) -> str:
        return self.retrieved_id(self.rank - 1) if self.rank else ""

    @property
    def matched_response(self) -> str:
        return self.retrieved_response(self.rank - 1) if self.rank else ""

    @property
    def other_suggestions(self) -> list[dict[str, str]]:
        """All retrieved candidates other than the matched one, in rank order."""
        matched_pos = self.rank - 1 if self.rank else None
        return [
            {"faq_id": str(doc.get(ID_KEY_NAME, "")), "response": str(doc.get(RETRIEVED_TEXT_KEY_NAME, ""))}
            for i, doc in enumerate(self.retrieved)
            if i != matched_pos
        ]


# --------------------------------------------------------------------------- #
# Response parsing
#
# The exact JSON shape returned by the graph endpoint was not verified
# against a live call (this environment's network policy blocks the host).
# `extract_retrieved_list` therefore tries a handful of known/likely key
# names first, then falls back to a recursive search for a list of dicts
# that carry the id_key_name field. Use --response-list-path to override
# once you've confirmed the real shape (see --probe-query), and --save-raw
# to keep every raw response for offline inspection.
# --------------------------------------------------------------------------- #

LIKELY_LIST_KEYS = [
    "documents", "retrieved_documents", "retrieval", "retrieved",
    "results", "data", "output", "response", "hits", "items",
]


def _get_by_dotted_path(obj: Any, dotted_path: str) -> Any:
    cur = obj
    for part in dotted_path.split("."):
        if isinstance(cur, dict):
            cur = cur.get(part)
        elif isinstance(cur, list) and part.isdigit():
            idx = int(part)
            cur = cur[idx] if idx < len(cur) else None
        else:
            return None
    return cur


def _find_id_list(obj: Any, id_key: str, _depth: int = 0) -> Optional[list[dict]]:
    """Recursively search for the first list of dicts that all contain id_key."""
    if _depth > 8:
        return None
    if isinstance(obj, list) and obj and all(isinstance(x, dict) for x in obj):
        if all(id_key in x for x in obj):
            return obj
    if isinstance(obj, dict):
        for key in LIKELY_LIST_KEYS:
            if key in obj:
                found = _find_id_list(obj[key], id_key, _depth + 1)
                if found is not None:
                    return found
        for v in obj.values():
            found = _find_id_list(v, id_key, _depth + 1)
            if found is not None:
                return found
    elif isinstance(obj, list):
        for v in obj:
            found = _find_id_list(v, id_key, _depth + 1)
            if found is not None:
                return found
    return None


def extract_retrieved_list(
    response_json: Any, id_key: str = ID_KEY_NAME, override_path: Optional[str] = None
) -> list[dict]:
    if override_path:
        candidate = _get_by_dotted_path(response_json, override_path)
        if isinstance(candidate, list):
            return candidate
        raise ValueError(
            f"--response-list-path {override_path!r} did not resolve to a list "
            f"(got {type(candidate).__name__}). Use --probe-query to inspect the real shape."
        )
    if isinstance(response_json, list):
        return response_json
    found = _find_id_list(response_json, id_key)
    if found is not None:
        return found
    return []


# --------------------------------------------------------------------------- #
# API call
# --------------------------------------------------------------------------- #

def build_request_body(
    query: str,
    path: str,
    lang: str,
    project_id: str,
    cms_project_id: str,
    graph_name: str,
) -> dict:
    return {
        "project_id": project_id,
        "cms_project_id": cms_project_id,
        "single_graph_name": graph_name,
        "path": path,
        "id_key_name": ID_KEY_NAME,
        "display_columns": DISPLAY_COLUMNS,
        "retrieved_text_key_name": RETRIEVED_TEXT_KEY_NAME,
        "query_lang": lang,
        "messages": [{"type": "human", "content": query}],
        "stream": False,
        "slim_response": True,
    }


def call_api(
    session: requests.Session,
    api_url: str,
    body: dict,
    headers: dict,
    timeout: float,
    retries: int,
    backoff_base: float = 1.5,
) -> tuple[Optional[dict], Optional[int], float, str]:
    """Returns (response_json, http_status, latency_ms, error_message)."""
    last_error = ""
    for attempt in range(retries + 1):
        start = time.monotonic()
        try:
            resp = session.post(api_url, json=body, headers=headers, timeout=timeout)
            latency_ms = (time.monotonic() - start) * 1000
            if resp.status_code >= 500 and attempt < retries:
                last_error = f"HTTP {resp.status_code} (retrying)"
                time.sleep(backoff_base ** attempt)
                continue
            try:
                data = resp.json()
            except ValueError:
                return None, resp.status_code, latency_ms, f"Non-JSON response: {resp.text[:300]!r}"
            if resp.status_code >= 400:
                return data, resp.status_code, latency_ms, f"HTTP {resp.status_code}: {resp.text[:300]}"
            return data, resp.status_code, latency_ms, ""
        except requests.RequestException as exc:
            latency_ms = (time.monotonic() - start) * 1000
            last_error = str(exc)
            if attempt < retries:
                time.sleep(backoff_base ** attempt)
                continue
            return None, None, latency_ms, last_error
    return None, None, 0.0, last_error


# --------------------------------------------------------------------------- #
# Test case loading
# --------------------------------------------------------------------------- #

def _pick_column(header_row: list[str], candidates: list[str]) -> Optional[int]:
    normalized = [h.strip() if isinstance(h, str) else h for h in header_row]
    for cand in candidates:
        if cand in normalized:
            return normalized.index(cand)
    return None


def _infer_lang_from_sheet_name(sheet_name: str, default_lang: str) -> str:
    """E.g. 'Benchmark Template cn' -> sc, via SHEET_NAME_LANG_SUFFIXES."""
    tokens = re.findall(r"[A-Za-z]+", sheet_name.lower())
    for token in reversed(tokens):
        if token in SHEET_NAME_LANG_SUFFIXES:
            return SHEET_NAME_LANG_SUFFIXES[token]
    return default_lang


def _resolve_path(raw_value: Any, default_path: Optional[str]) -> Optional[str]:
    if raw_value is None or str(raw_value).strip() == "":
        return default_path
    key = str(raw_value).strip()
    if key in VALID_PATHS:
        return key
    return CATEGORY_TO_PATH.get(key.lower().replace(" ", "_"), key)


def _load_sheet(
    wb, sheet_name: str, source_path: Path, default_path: Optional[str], default_lang: str
) -> list[TestCase]:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = list(rows[0])

    query_idx = _pick_column(header, TESTCASE_QUERY_COLS)
    expected_idx = _pick_column(header, TESTCASE_EXPECTED_COLS)
    path_idx = _pick_column(header, TESTCASE_PATH_COLS)
    lang_idx = _pick_column(header, TESTCASE_LANG_COLS)
    notes_idx = _pick_column(header, TESTCASE_NOTES_COLS)

    if query_idx is None or expected_idx is None:
        return []

    cases: list[TestCase] = []
    for row_num, row in enumerate(rows[1:], start=2):
        query = row[query_idx] if query_idx < len(row) else None
        if query is None or str(query).strip() == "":
            continue
        query_str = str(query).strip()
        if EG_PREFIX_RE.match(query_str):
            continue

        expected_raw = row[expected_idx] if expected_idx < len(row) else None
        expected_ids = [e.strip() for e in str(expected_raw or "").split(",") if e.strip()]
        if not expected_ids:
            raise ValueError(
                f"Row {row_num} in sheet {sheet_name!r} of {source_path} has a query but no "
                f"expected FAQ ID."
            )

        raw_path = row[path_idx] if path_idx is not None and path_idx < len(row) else None
        row_path = _resolve_path(raw_path, default_path)
        if not row_path:
            raise ValueError(
                f"Row {row_num} in sheet {sheet_name!r} of {source_path} has no Path/Category "
                f"and no --default-path was given."
            )
        if row_path not in VALID_PATHS:
            raise ValueError(
                f"Row {row_num} in sheet {sheet_name!r} of {source_path} has Path/Category="
                f"{raw_path!r}, which doesn't map to one of {VALID_PATHS}. Add a mapping in "
                f"CATEGORY_TO_PATH or fix the sheet."
            )

        row_lang = None
        if lang_idx is not None and lang_idx < len(row):
            row_lang = row[lang_idx]
        row_lang = str(row_lang).strip() if row_lang else default_lang

        row_notes = ""
        if notes_idx is not None and notes_idx < len(row):
            row_notes = str(row[notes_idx] or "")

        cases.append(
            TestCase(
                row_num=row_num,
                query=query_str,
                expected_ids=expected_ids,
                path=row_path,
                lang=row_lang,
                notes=row_notes,
                sheet=sheet_name,
            )
        )
    return cases


def load_testcases(path: Path, default_path: Optional[str], default_lang: str) -> list[TestCase]:
    """
    Supports two workbook layouts:

    1. A single "Test Cases" sheet (see sino_benchmark_testcases_TEMPLATE.xlsx):
       Query / Expected_FAQ_ID / Path / Language / Notes columns.

    2. Sino's own multi-language benchmark template: one sheet per language
       (e.g. "Benchmark Template zh/en/cn") with No. / Testing Query /
       Category / FAQ Reference No. columns. query_lang is inferred per
       sheet from its name (see SHEET_NAME_LANG_SUFFIXES), Category is
       mapped to a Path via CATEGORY_TO_PATH, and "EG:"-prefixed example
       rows are skipped automatically. All sheets with a recognizable
       header are combined into one test run.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    if "Test Cases" in wb.sheetnames:
        return _load_sheet(wb, "Test Cases", path, default_path, default_lang)

    cases: list[TestCase] = []
    for sheet_name in wb.sheetnames:
        sheet_lang = _infer_lang_from_sheet_name(sheet_name, default_lang)
        cases.extend(_load_sheet(wb, sheet_name, path, default_path, sheet_lang))

    if not cases:
        raise ValueError(
            f"Could not find a usable test-case sheet in {path}. Expected either a sheet named "
            f"'Test Cases', or one or more sheets with a query column (one of "
            f"{TESTCASE_QUERY_COLS}) and an expected-answer column (one of "
            f"{TESTCASE_EXPECTED_COLS}). Sheets found: {wb.sheetnames}"
        )
    return cases


# --------------------------------------------------------------------------- #
# Scoring
# --------------------------------------------------------------------------- #

def score_result(case: TestCase, retrieved: list[dict]) -> Optional[int]:
    expected_set = {e.lower() for e in case.expected_ids}
    for i, doc in enumerate(retrieved):
        doc_id = str(doc.get(ID_KEY_NAME, "")).strip().lower()
        if doc_id in expected_set:
            return i + 1
    return None


# --------------------------------------------------------------------------- #
# Output workbook
# --------------------------------------------------------------------------- #

def write_results(
    results: list[TestResult],
    output_path: Path,
    top_ks: list[int],
    max_retrieved_cols: int = 5,
    save_raw: bool = False,
):
    wb = openpyxl.Workbook()

    # ---- Details sheet ----
    ws = wb.active
    ws.title = "Details"
    header = [
        "Sheet", "Row", "Path", "Language", "Query", "Expected_FAQ_ID",
        "Top1_Correct", "Rank_of_Expected", "Reciprocal_Rank",
    ]
    header += [f"Hit@{k}" for k in top_ks]
    header += ["Matched_FAQ_ID", "Matched_Response"]
    for i in range(1, max_retrieved_cols + 1):
        header += [f"Suggested_{i}_FAQ_ID", f"Suggested_{i}_Response"]
    header += ["Latency_ms", "HTTP_Status", "Error", "Notes"]
    if save_raw:
        header.append("Raw_Response")
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    for r in results:
        row = [
            r.case.sheet, r.case.row_num, r.case.path, r.case.lang, r.case.query,
            ", ".join(r.case.expected_ids),
            "Y" if r.top1_correct else "N",
            r.rank if r.rank else "",
            round(r.reciprocal_rank, 4),
        ]
        row += ["Y" if r.hit_at(k) else "N" for k in top_ks]
        row += [r.matched_id, r.matched_response]
        for i in range(max_retrieved_cols):
            row += [r.retrieved_id(i), r.retrieved_response(i)]
        row += [
            round(r.latency_ms, 1) if r.latency_ms is not None else "",
            r.http_status if r.http_status is not None else "",
            r.error,
            r.case.notes,
        ]
        if save_raw:
            row.append(json.dumps(r.raw_response, ensure_ascii=False)[:32000])
        ws.append(row)

    col_widths = (
        [20, 6, 14, 10, 45, 16]  # Sheet, Row, Path, Language, Query, Expected_FAQ_ID
        + [12, 16, 16]  # Top1_Correct, Rank_of_Expected, Reciprocal_Rank
        + [8] * len(top_ks)  # Hit@k...
        + [16, 45]  # Matched_FAQ_ID, Matched_Response
        + [16, 45] * max_retrieved_cols  # Suggested_i_FAQ_ID, Suggested_i_Response
        + [12, 12, 40, 25]  # Latency_ms, HTTP_Status, Error, Notes
    )
    for i, col_width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = col_width

    # ---- Summary sheet ----
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Overall"] + list(VALID_PATHS))
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    def subset(path: Optional[str]):
        return [r for r in results if path is None or r.case.path == path]

    groups = [None] + list(VALID_PATHS)

    def metric_row(label, fn):
        ws2.append([label] + [fn(subset(g)) for g in groups])

    def pct(n, d):
        return round(100.0 * n / d, 1) if d else 0.0

    metric_row("Test Cases", lambda rs: len(rs))
    metric_row("Errors", lambda rs: sum(1 for r in rs if r.error))
    metric_row("Scored (no error)", lambda rs: sum(1 for r in rs if not r.error))
    metric_row(
        "Top-1 Accuracy (%)",
        lambda rs: pct(sum(1 for r in rs if not r.error and r.top1_correct), sum(1 for r in rs if not r.error)),
    )
    for k in top_ks:
        metric_row(
            f"Recall@{k} (%)",
            lambda rs, k=k: pct(sum(1 for r in rs if not r.error and r.hit_at(k)), sum(1 for r in rs if not r.error)),
        )
    metric_row(
        "MRR",
        lambda rs: round(
            sum(r.reciprocal_rank for r in rs if not r.error) / max(1, sum(1 for r in rs if not r.error)), 4
        ),
    )
    metric_row(
        "Avg Latency (ms)",
        lambda rs: round(
            sum(r.latency_ms for r in rs if r.latency_ms is not None)
            / max(1, sum(1 for r in rs if r.latency_ms is not None)),
            1,
        ),
    )
    languages = sorted({r.case.lang for r in results})
    if len(languages) > 1:
        ws2.append([])
        ws2.append(["By Language", "Test Cases", "Errors", "Top-1 Accuracy (%)", "MRR"])
        for cell in ws2[ws2.max_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        for lang in languages:
            lang_results = [r for r in results if r.case.lang == lang]
            lang_scored = [r for r in lang_results if not r.error]
            ws2.append(
                [
                    lang,
                    len(lang_results),
                    len(lang_results) - len(lang_scored),
                    pct(sum(1 for r in lang_scored if r.top1_correct), len(lang_scored)),
                    round(sum(r.reciprocal_rank for r in lang_scored) / max(1, len(lang_scored)), 4),
                ]
            )

    ws2.append([])
    ws2.append([f"Generated: {datetime.now().isoformat(timespec='seconds')}"])

    for i, w in enumerate([24, 12, 16, 16], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #

def parse_headers(header_args: list[str]) -> dict:
    headers = {}
    for h in header_args or []:
        if ":" not in h:
            raise ValueError(f"--header value {h!r} must be in 'Key: Value' form.")
        key, _, value = h.partition(":")
        headers[key.strip()] = value.strip()
    return headers


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--testcases", type=Path, help="Input .xlsx of test cases.")
    ap.add_argument("--output", type=Path, help="Output .xlsx path (default: results_<timestamp>.xlsx).")
    ap.add_argument("--api-url", default=DEFAULT_API_URL)
    ap.add_argument("--project-id", default=DEFAULT_PROJECT_ID)
    ap.add_argument("--cms-project-id", default=DEFAULT_CMS_PROJECT_ID)
    ap.add_argument("--graph-name", default=DEFAULT_GRAPH_NAME)
    ap.add_argument("--default-path", choices=VALID_PATHS, help="Path to use for rows without one.")
    ap.add_argument("--default-lang", default="en", help="query_lang to use for rows without one.")
    ap.add_argument("--top-k", default="1,3,5", help="Comma-separated k values for Recall@k, e.g. '1,3,5'.")
    ap.add_argument("--timeout", type=float, default=30.0)
    ap.add_argument("--delay", type=float, default=0.2, help="Seconds to sleep between requests.")
    ap.add_argument("--retries", type=int, default=2, help="Retries on network error / HTTP 5xx.")
    ap.add_argument("--header", action="append", help="Extra request header 'Key: Value'. Repeatable.")
    ap.add_argument("--response-list-path", help="Dotted path to the retrieved-documents list in the response, e.g. 'data.documents'. Overrides auto-detection.")
    ap.add_argument("--max-retrieved-cols", type=int, default=5, help="How many retrieved ranks to include as columns.")
    ap.add_argument("--limit", type=int, help="Only run the first N test cases (smoke test).")
    ap.add_argument("--save-raw", action="store_true", help="Include the full raw JSON response per row in the output.")
    ap.add_argument("--verbose", action="store_true")
    ap.add_argument("--probe-query", help="Send a single ad-hoc query and pretty-print the raw response, then exit.")
    ap.add_argument("--probe-path", choices=VALID_PATHS, help="Path to use with --probe-query.")
    ap.add_argument("--probe-lang", default="en", help="query_lang to use with --probe-query.")
    args = ap.parse_args()

    headers = parse_headers(args.header)
    session = requests.Session()

    if args.probe_query:
        if not args.probe_path:
            ap.error("--probe-path is required with --probe-query.")
        body = build_request_body(
            args.probe_query, args.probe_path, args.probe_lang,
            args.project_id, args.cms_project_id, args.graph_name,
        )
        print("Request body:")
        print(json.dumps(body, indent=2, ensure_ascii=False))
        data, status, latency_ms, error = call_api(
            session, args.api_url, body, headers, args.timeout, args.retries
        )
        print(f"\nHTTP {status}  ({latency_ms:.0f} ms)")
        if error:
            print(f"Error: {error}")
        print("\nRaw response:")
        print(json.dumps(data, indent=2, ensure_ascii=False) if data is not None else "<no JSON body>")
        if data is not None:
            found = extract_retrieved_list(data, ID_KEY_NAME, args.response_list_path)
            print(f"\nAuto-detected retrieved list ({len(found)} item(s)):")
            print(json.dumps(found, indent=2, ensure_ascii=False))
            if not found:
                print(
                    "\nNo list containing an 'FAQ_ID' field was found automatically. "
                    "Inspect the raw response above and pass --response-list-path "
                    "(e.g. --response-list-path data.documents) when running the full benchmark."
                )
        return

    if not args.testcases:
        ap.error("--testcases is required (or use --probe-query to inspect the API first).")

    top_ks = sorted({int(k) for k in args.top_k.split(",") if k.strip()})
    output_path = args.output or Path(f"results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    print(f"Loading test cases from {args.testcases} ...")
    cases = load_testcases(args.testcases, args.default_path, args.default_lang)
    if args.limit:
        cases = cases[: args.limit]
    print(f"Loaded {len(cases)} test case(s).")

    results: list[TestResult] = []
    for i, case in enumerate(cases, start=1):
        body = build_request_body(
            case.query, case.path, case.lang,
            args.project_id, args.cms_project_id, args.graph_name,
        )
        data, status, latency_ms, error = call_api(
            session, args.api_url, body, headers, args.timeout, args.retries
        )
        result = TestResult(case=case, http_status=status, latency_ms=latency_ms, error=error, raw_response=data)
        if data is not None and not error:
            try:
                retrieved = extract_retrieved_list(data, ID_KEY_NAME, args.response_list_path)
            except ValueError as exc:
                retrieved = []
                result.error = str(exc)
            result.retrieved = retrieved
            result.rank = score_result(case, retrieved)
        if result.error:
            print(f"[{i}/{len(cases)}] row {case.row_num} ({case.path}): ERROR: {result.error}")
        elif args.verbose:
            print(f"[{i}/{len(cases)}] row {case.row_num} ({case.path}): "
                  f"rank={result.rank} top1={'Y' if result.top1_correct else 'N'} "
                  f"({result.latency_ms:.0f} ms)")
        results.append(result)
        if args.delay and i < len(cases):
            time.sleep(args.delay)

    write_results(results, output_path, top_ks, args.max_retrieved_cols, args.save_raw)
    print(f"\nWrote results to {output_path}")

    scored = [r for r in results if not r.error]
    if scored:
        top1 = 100.0 * sum(1 for r in scored if r.top1_correct) / len(scored)
        mrr = sum(r.reciprocal_rank for r in scored) / len(scored)
        print(f"Top-1 accuracy: {top1:.1f}%   MRR: {mrr:.4f}   "
              f"(scored {len(scored)}/{len(results)}, {len(results) - len(scored)} error(s))")


if __name__ == "__main__":
    main()
